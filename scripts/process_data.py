from __future__ import annotations

import argparse
import csv
import io
import json
import re
import sys
import unicodedata
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_MIN_ANALYSIS_SPEED = 50
DEFAULT_SPEED_LIMIT = 80
SUPPORTED_INPUT_EXTENSIONS = {".csv", ".xlsx", ".xlsm"}

# Mapeamento flexivel para aceitar pequenas variacoes de cabecalho nas planilhas.
HEADER_ALIASES = {
    "motorista": "vehicle",
    "veiculo": "vehicle",
    "placa": "plate",
    "placa_": "plate",
    "placa_do_veiculo": "plate",
    "data_hora": "datetime",
    "datahora": "datetime",
    "data_e_horario_de_inicio_do_evento_de_velocidade": "datetime",
    "velocidade": "speed",
    "media_da_velocidade_durante_o_evento_km_h": "speed",
    "evento": "event",
    "eventos": "event",
    "endereco": "address",
    "gravidade": "severity",
    "grupo": "group",
    "limite_de_velocidade_da_via_km_h": "road_speed_limit",
    "tempo_total_acima_do_limite_segundos": "event_duration_seconds",
}

DATETIME_PATTERNS = (
    "%d/%m/%Y %H:%M:%S",
    "%d/%m/%Y %H:%M",
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
)


def main() -> int:
    args = parse_args()
    input_dirs = [path.resolve() for path in args.input_dir]
    output_json = args.output_json.resolve()
    output_js = args.output_js.resolve()

    files = collect_input_files(input_dirs)

    if not files:
        searched_dirs = ", ".join(str(path) for path in input_dirs)
        print(
            f"Nenhum arquivo encontrado em {searched_dirs}. "
            "Adicione planilhas .csv, .xlsx ou .xlsm e rode novamente.",
            file=sys.stderr,
        )
        return 1

    records = []

    for source_file in files:
        for raw_row in read_source_rows(source_file):
            records.append(
                normalize_row(
                    raw_row=raw_row,
                    source_file=source_file.name,
                    speed_limit=args.speed_limit,
                )
            )

    records.sort(
        key=lambda item: (
            item["timestampSortable"],
            item["sourceFile"],
            item["rowNumber"],
        ),
        reverse=True,
    )

    payload = {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "sourceFiles": [file.name for file in files],
        "config": {
            "defaultMinAnalysisSpeed": args.min_analysis_speed,
            "defaultSpeedLimit": args.speed_limit,
            "usesPerEventSpeedLimit": any(
                record.get("roadSpeedLimit") is not None for record in records
            ),
        },
        "summary": build_summary(records, args.speed_limit),
        "records": records,
    }

    write_json(output_json, payload)
    write_js(output_js, payload)

    print("Arquivos gerados com sucesso:")
    print(f"- JSON: {output_json}")
    print(f"- JS:   {output_js}")
    print(f"- Registros: {len(records)}")
    print(f"- Alertas no limite padrao: {payload['summary']['defaultAlertCount']}")
    return 0


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Converte planilhas de velocidade em arquivos prontos para o dashboard estatico."
    )
    parser.add_argument(
        "--input-dir",
        type=Path,
        action="append",
        default=None,
        help="Pasta com arquivos .csv, .xlsx ou .xlsm. Pode ser informada mais de uma vez.",
    )
    parser.add_argument(
        "--output-json",
        type=Path,
        default=Path("data/dashboard-data.json"),
        help="Arquivo JSON final usado pelo dashboard.",
    )
    parser.add_argument(
        "--output-js",
        type=Path,
        default=Path("data/dashboard-data.js"),
        help="Arquivo JS espelho para compatibilidade ao abrir index.html localmente.",
    )
    parser.add_argument(
        "--min-analysis-speed",
        type=int,
        default=DEFAULT_MIN_ANALYSIS_SPEED,
        help="Velocidade minima de analise salva como padrao no dashboard.",
    )
    parser.add_argument(
        "--speed-limit",
        type=int,
        default=DEFAULT_SPEED_LIMIT,
        help="Limite de velocidade padrao para alertas e severidade.",
    )
    args = parser.parse_args()

    if not args.input_dir:
        args.input_dir = [Path("."), Path("input")]

    return args


def collect_input_files(input_dirs: list[Path]) -> list[Path]:
    discovered_files: dict[str, Path] = {}

    for input_dir in input_dirs:
        if not input_dir.exists():
            continue

        for path in sorted(input_dir.iterdir()):
            if (
                path.is_file()
                and path.suffix.lower() in SUPPORTED_INPUT_EXTENSIONS
                and not path.name.startswith("~$")
            ):
                discovered_files[str(path.resolve()).lower()] = path.resolve()

    return sorted(discovered_files.values(), key=lambda path: path.name.lower())


def read_source_rows(path: Path) -> list[dict]:
    if path.suffix.lower() == ".csv":
        return read_csv_rows(path)

    return read_excel_rows(path)


def read_csv_rows(path: Path) -> list[dict]:
    text = read_text_with_fallback(path)
    sample = text[:4096]

    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,")
    except csv.Error:
        dialect = csv.excel
        dialect.delimiter = ";"

    rows = []
    stream = io.StringIO(text)
    reader = csv.DictReader(stream, dialect=dialect)

    for row_number, row in enumerate(reader, start=2):
        if row is None or not any(clean_text(value) for value in row.values()):
            continue

        rows.append(
            {
                "sheetName": "CSV",
                "rowNumber": row_number,
                "values": canonicalize_row(row),
            }
        )

    return rows


def read_excel_rows(path: Path) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows = []

    for worksheet in workbook.worksheets:
        iterator = worksheet.iter_rows(values_only=True)
        header_row = None
        header_index = 0

        for row_index, row in enumerate(iterator, start=1):
            if any(clean_text(value) for value in row):
                header_row = row
                header_index = row_index
                break

        if not header_row:
            continue

        headers = [canonicalize_header(value) for value in header_row]

        for row_index, row in enumerate(iterator, start=header_index + 1):
            if row is None or not any(clean_text(value) for value in row):
                continue

            values = {}

            for index, header in enumerate(headers):
                if not header:
                    continue

                values[header] = row[index] if index < len(row) else None

            rows.append(
                {
                    "sheetName": worksheet.title,
                    "rowNumber": row_index,
                    "values": values,
                }
            )

    workbook.close()
    return rows


def canonicalize_row(row: dict) -> dict:
    values = {}

    for key, value in row.items():
        header = canonicalize_header(key)
        if header:
            values[header] = value

    return values


def canonicalize_header(value: object) -> str:
    text = clean_text(value)

    if not text:
        return ""

    ascii_text = (
        unicodedata.normalize("NFD", text)
        .encode("ascii", "ignore")
        .decode("ascii")
        .lower()
    )
    ascii_text = re.sub(r"[^a-z0-9]+", "_", ascii_text).strip("_")
    return HEADER_ALIASES.get(ascii_text, ascii_text)


def normalize_row(raw_row: dict, source_file: str, speed_limit: int) -> dict:
    values = raw_row["values"]

    plate = normalize_plate(values.get("plate"))
    vehicle = clean_text(values.get("vehicle"))
    group = normalize_label(values.get("group"), fallback="Sem grupo")
    source_severity = normalize_source_severity(values.get("severity"))
    event = normalize_label(values.get("event"), fallback="Excesso de velocidade")
    address = normalize_label(
        values.get("address"),
        fallback=group if group and group != "Sem grupo" else "Endereco nao informado",
    )

    parsed_datetime = parse_datetime(values.get("datetime"))
    speed = parse_speed(values.get("speed"))
    road_speed_limit = parse_speed(values.get("road_speed_limit"))
    event_duration_seconds = parse_speed(values.get("event_duration_seconds"))

    validation_issues = []

    if parsed_datetime is None:
        validation_issues.append("data_hora_invalida")

    if speed is None:
        validation_issues.append("velocidade_invalida")

    is_valid_record = parsed_datetime is not None and speed is not None
    effective_speed_limit = road_speed_limit if road_speed_limit is not None else speed_limit
    severity = source_severity or classify_severity(
        speed,
        effective_speed_limit,
        is_valid_record,
    )

    timestamp_sortable = parsed_datetime.strftime("%Y-%m-%dT%H:%M:%S") if parsed_datetime else ""
    display_datetime = parsed_datetime.strftime("%d/%m/%Y %H:%M:%S") if parsed_datetime else "Data invalida"
    date_key = parsed_datetime.strftime("%Y-%m-%d") if parsed_datetime else ""

    return {
        "id": f"{Path(source_file).stem}-{raw_row['sheetName']}-{raw_row['rowNumber']}",
        "sourceFile": source_file,
        "sourceSheet": raw_row["sheetName"],
        "rowNumber": raw_row["rowNumber"],
        "vehicle": vehicle,
        "plate": plate,
        "timestamp": timestamp_sortable,
        "timestampSortable": timestamp_sortable,
        "displayDateTime": display_datetime,
        "dateKey": date_key,
        "speed": speed,
        "speedDisplay": format_speed(speed),
        "event": event,
        "address": address,
        "group": group,
        "sourceSeverity": source_severity,
        "roadSpeedLimit": road_speed_limit,
        "effectiveSpeedLimit": effective_speed_limit if is_valid_record else road_speed_limit,
        "eventDurationSeconds": event_duration_seconds,
        "severityDefault": severity,
        "isAlertDefault": bool(is_valid_record and speed > effective_speed_limit),
        "isValidDate": parsed_datetime is not None,
        "isValidSpeed": speed is not None,
        "isValidRecord": is_valid_record,
        "validationIssues": validation_issues,
    }


def build_summary(records: list[dict], speed_limit: int) -> dict:
    valid_records = [record for record in records if record["isValidRecord"]]
    unique_plates = {
        record["plate"]
        for record in records
        if record["plate"] and record["plate"] != "Nao informado"
    }

    return {
        "totalRecords": len(records),
        "validRecords": len(valid_records),
        "invalidRecords": len(records) - len(valid_records),
        "defaultAlertCount": sum(1 for record in records if record["isAlertDefault"]),
        "maxSpeed": max((record["speed"] for record in valid_records), default=0),
        "monitoredPlates": len(unique_plates),
    }


def read_text_with_fallback(path: Path) -> str:
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue

    return path.read_text(encoding="utf-8", errors="replace")


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_label(value: object, fallback: str) -> str:
    text = clean_text(value)
    return text if text else fallback


def normalize_plate(value: object) -> str:
    text = clean_text(value).upper()
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"[^A-Z0-9-]", "", text)
    text = re.sub(r"-{2,}", "-", text).strip("-")
    return text or "Nao informado"


def parse_datetime(value: object) -> datetime | None:
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.replace(microsecond=0)

    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())

    text = clean_text(value)

    if not text:
        return None

    for pattern in DATETIME_PATTERNS:
        try:
            return datetime.strptime(text, pattern)
        except ValueError:
            continue

    return None


def parse_speed(value: object) -> int | float | None:
    if value is None:
        return None

    if isinstance(value, (int, float)):
        return compact_number(float(value))

    text = clean_text(value)

    if not text:
        return None

    match = re.search(r"-?\d+(?:[.,]\d+)?", text)

    if not match:
        return None

    numeric_text = match.group(0).replace(",", ".")

    try:
        return compact_number(float(numeric_text))
    except ValueError:
        return None


def compact_number(value: float) -> int | float:
    if value.is_integer():
        return int(value)
    return round(value, 1)


def normalize_source_severity(value: object) -> str:
    text = clean_text(value)

    if not text:
        return ""

    ascii_text = (
        unicodedata.normalize("NFD", text)
        .encode("ascii", "ignore")
        .decode("ascii")
        .lower()
    )

    severity_map = {
        "leve": "Normal",
        "media": "Atencao",
        "grave": "Critico",
        "gravissima": "Critico",
    }

    return severity_map.get(ascii_text, "")


def classify_severity(speed: int | float | None, speed_limit: int, is_valid_record: bool) -> str:
    if not is_valid_record or speed is None:
        return "Dados invalidos"

    if speed <= speed_limit:
        return "Normal"

    if speed <= speed_limit + 10:
        return "Atencao"

    return "Critico"


def format_speed(speed: int | float | None) -> str:
    if speed is None:
        return "Nao informado"
    return f"{speed} km/h"


def write_json(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def write_js(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    content = (
        "// Arquivo gerado automaticamente. Nao edite manualmente.\n"
        "window.FLEET_SPEED_DASHBOARD_DATA = "
        f"{json.dumps(payload, ensure_ascii=False, indent=2)};\n"
    )
    path.write_text(content, encoding="utf-8")


if __name__ == "__main__":
    raise SystemExit(main())
